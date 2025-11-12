'''
This script allow to extract Property Values defined with PVMT associated to elements (in this case SystemFunctions)
'''
# To run it:
#  - enable Developer capabilities if not already done (see documentation in the help menu)
#  - you can run this script by launching the contextual menu "Run As / EASE Script..." 
#    on this script. 
#    - By default, the model selected is IFE sample (aird path of the model written below)

#  - you can also run this script according to a configuration (script selected, arguments) 
#    and modify the configuration by launching the contextual menu "Run As / Run configurations..." 
#    on this script. 
#    - create a new "EASE Script" configuration
#    - define the name of the configuration: "Export_Property_Values_associated_to_elements_to_xlsx.py" (for instance)
#    - define the Script Source path: "workspace://Python4Capella/sample_scripts/Export_Property_Values_associated_to_elements_to_xlsx.py"

# include needed for the Capella modeller API
include('workspace://Python4Capella/simplified_api/capella.py')
if False:
    from simplified_api.capella import *
    
# include needed for the PVMT API
include('workspace://Python4Capella/simplified_api/pvmt.py')
if False:
    from simplified_api.pvmt import *
    
# include needed for utilities
include('workspace://Python4Capella/utilities/CapellaPlatform.py')
if False:
    from utilities.CapellaPlatform import *
    
# include needed to read/write xlsx files
from openpyxl import *

# change this path to execute the script on your model (here is the IFE sample). 
# comment it if you want to use the "Run configuration" instead
aird_path = '/In-Flight Entertainment System/In-Flight Entertainment System.aird'

'''
#Here is the "Run Configuration" part to uncomment if you want to use this functionality :

#check parameter numbers
if len(argv) != 1:
    # use IFE default values
    aird_path = "/In-Flight Entertainment System/In-Flight Entertainment System.aird"
else:
    # Load the Capella model from the first argument of the script
    aird_path = argv[0]
'''

model = CapellaModel()
model.open(aird_path)

# gets the SystemEngineering and print its name
se = model.get_system_engineering()
print('starting export of model ' + se.get_name())

# preparing excel file export
project_name = aird_path[0:(aird_path.index("/", 1) + 1)]
project = CapellaPlatform.getProject(project_name)
folder = CapellaPlatform.getFolder(project, 'results')
CapellaPlatform.getAbsolutePath(folder)
xlsx_file_name = CapellaPlatform.getAbsolutePath(folder) + '/' + 'Export_Property_Values_associated_to_elements_to_xlsx.xlsx'
# create a workbook
workbook = Workbook()


# writing excel file header
worksheet = workbook.active
worksheet.title = 'PV export on Functions'
worksheet["A1"] = 'System Function name'

# retrieve the list of PV to write the header
allPVs = []

# change SystemFunction by another type to retrieve PV for other elements
allSF = se.get_all_contents_by_type(SystemFunction)

for sf in allSF:
    for pvName in PVMT.get_p_v_names(sf):
        if pvName not in allPVs:
            allPVs.append(pvName)

# writing header for all PV found
j = 2
for pvName in allPVs:
    worksheet.cell(row = 1, column = j).value = pvName
    j = j + 1

# now retrieving elements from model
i = 2

for sf in allSF:
    worksheet.cell(row = i, column = 1).value = sf.get_name()
    j = 2
    for pvName in allPVs:
        worksheet.cell(row = i, column = j).value = PVMT.get_p_v_value(sf, pvName)
        j = j + 1
    i = i + 1

# Save the xlsx file
workbook.save(xlsx_file_name)

print('saving excel file')

# refresh 
CapellaPlatform.refresh(folder)
