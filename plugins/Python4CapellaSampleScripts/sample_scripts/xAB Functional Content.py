# name                 : Extract content of xAB diagram
# script-type          : Python
# description          : Extract content of xAB diagram
# popup                : enableFor(org.eclipse.sirius.viewpoint.DRepresentationDescriptor)

'''
This script allows to extract content of xAB diagram
It will create a folder result in the selected Capella project with the resulting xlsx file.
'''

# To run it:
#  - you need to right click on a DRepresentation Descriptor (diagram) and select the "Extract content of xAB diagram" menu
#  - enable Developer capabilities if not already done (see documentation in the help menu)
#  - you can run this script by launching the contextual menu "Run As / EASE Script..." 
#    on this script. 
#    - By default, the model selected is IFE sample (aird path of the model written below)

#  - you can also run this script according to a configuration (script selected, arguments) 
#    and modify the configuration by launching the contextual menu "Run As / Run configurations..." 
#    on this script. 
#    - create a new "EASE Script" configuration
#    - define the name of the configuration: "Export_xAB_Functional_Content.py" (for instance)
#    - define the Script Source path: "workspace://Python4Capella/sample_scripts/Export_xAB_Functional_Content.py"
#


# include needed for the Capella modeller API
include('workspace://Python4Capella/simplified_api/capella.py')
if False:
    from simplified_api.capella import *
    
# include needed for the PVMT API
include('workspace://Python4Capella/simplified_api/pvmt.py')
if False:
    from simplified_api.pvmt import *
    
 # include needed for utilities
include('workspace://Python4Capella/utilities/CapellaPlatform.py')
if False:
    from utilities.CapellaPlatform import *
    
# include needed to read/write xlsx files
from openpyxl import *
from openpyxl.formatting.rule import ColorScaleRule,CellIsRule,FormulaRule
from openpyxl.styles import Alignment, NamedStyle, Font, colors, Color, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles.borders import Border, Side

# Retrieve the Element from the current selection and its aird model path
myDiagram = Diagram(CapellaPlatform.getFirstSelectedElement())
aird_path = '/In-Flight Entertainment System/In-Flight Entertainment System.aird'

'''
# change this path to execute the script on your model (here is the IFE sample). 
# Uncomment it if you want to use the "Run configuration" instead
aird_path = '/In-Flight Entertainment System/In-Flight Entertainment System.aird'
'''
'''
#Here is the "Run Configuration" part to uncomment if you want to use this functionality :
#check parameter numbers
if len(argv) != 1:
    # use IFE default values
    aird_path = "/In-Flight Entertainment System/In-Flight Entertainment System.aird"
else:
    # Load the Capella model from the first argument of the script
    aird_path = argv[0]
'''

model = CapellaModel()
model.open(aird_path)

# gets the SystemEngineering and print its name
se = model.get_system_engineering()
system = se.get_system_analysis().get_system()
internal_functions = []
external_functions = []
component_exchanges = []
system_components = []
actors = []

print('starting export of diagram ' + myDiagram.get_name())

for element in myDiagram.get_represented_elements():
    if isinstance(element, SystemFunction):
        if not element.get_all_contents_by_type(SystemFunction):
            if element.get_allocating_component().get_java_object().isActor() == True:
                external_functions.append(element)
            else:
                internal_functions.append(element)
        else:
            print(element.get_name() + ' is not a leaf function. It internal/external status was not computed')

for element in myDiagram.get_represented_elements():
    if isinstance(element, LogicalFunction):
        if not element.get_all_contents_by_type(LogicalFunction):
            if element.get_allocating_component().get_java_object().isActor() == True:
                external_functions.append(element)
            else:
                internal_functions.append(element)
        else:
            print(element.get_name() + ' is not a leaf function. It internal/external status was not computed')

for element in myDiagram.get_represented_elements():
    if isinstance(element, ComponentExchange):
        component_exchanges.append(element)
        
for element in myDiagram.get_represented_elements():
    if isinstance(element, System) or isinstance(element, LogicalComponent):
        system_components.append(element)
            
for element in myDiagram.get_represented_elements():
    if isinstance(element, SystemActor) or isinstance(element, LogicalActor):
        actors.append(element)
      
print('----- ACTORS -----')
for act in actors:
    print('--> ' + act.get_name())
print('----- COMPONENTS -----')
for sc in system_components:
    print('--> ' + sc.get_name())
print('----- EXTERNAL FUNCTIONS -----')
for e_f in external_functions:
    print('--> ' + e_f.get_name())
print('----- INTERNAL FUNCTIONS -----')
for i_f in internal_functions:
    print('--> ' + i_f.get_name())
print('----- COMPONENT EXCHANGES -----')
for ce in component_exchanges:
    print('--> ' + ce.get_name())

# preparing excel file export
project_name = aird_path[0:(aird_path.index("/", 1) + 1)]
project = CapellaPlatform.getProject(project_name)
folder = CapellaPlatform.getFolder(project, 'results')
xlsx_file_name = CapellaPlatform.getAbsolutePath(folder) + '/' + 'Diagram Content - ' + myDiagram.get_name() + '.xlsx'

# create a workbook
workbook = Workbook()

# writing excel file header
worksheet1 = workbook.active
worksheet1.title = 'Diagram Content'

i = 1
worksheet1.cell(row = i, column = 1).value = myDiagram.get_name()
worksheet1.merge_cells(start_row=1, start_column=1, end_row =1, end_column=2)


index_first_external_functions = i + 1
for e_f in external_functions:
    i=i+1 
    worksheet1.cell(row = i, column = 1).value = 'External Functions'
    worksheet1.cell(row = i, column = 2).value = e_f.get_name()
    
worksheet1.merge_cells(start_row=index_first_external_functions, start_column=1, end_row =i, end_column=1)

index_first_internal_functions = i+1
for i_f in internal_functions:
    i=i+1 
    worksheet1.cell(row = i, column = 1).value = 'Internal Functions'
    worksheet1.cell(row = i, column = 2).value = i_f.get_name()
    
worksheet1.merge_cells(start_row=index_first_internal_functions, start_column=1, end_row =i, end_column=1)

index_first_actors = i + 1
for act in actors:
    i=i+1 
    worksheet1.cell(row = i, column = 1).value = 'Actors'
    worksheet1.cell(row = i, column = 2).value = act.get_name()
    
worksheet1.merge_cells(start_row=index_first_actors, start_column=1, end_row =i, end_column=1)

index_first_system_components = i + 1
for s_c in system_components:
    i=i+1 
    worksheet1.cell(row = i, column = 1).value = 'System/System Components'
    worksheet1.cell(row = i, column = 2).value = s_c.get_name()
    
worksheet1.merge_cells(start_row=index_first_system_components, start_column=1, end_row =i, end_column=1)

index_first_component_exchanges = i + 1
for c_e in component_exchanges:
    i=i+1 
    worksheet1.cell(row = i, column = 1).value = 'Interfaces'
    worksheet1.cell(row = i, column = 2).value = c_e.get_name()
    
worksheet1.merge_cells(start_row=index_first_component_exchanges, start_column=1, end_row =i, end_column=1)
  
column_widths = [] 
bd = Side(border_style='medium')
chosen_border = Border(left=bd, top=bd,right=bd, bottom=bd)
chosen_alignment = Alignment(wrap_text=True,vertical='top')

for row in worksheet1.iter_rows():
    for cell in row:
        cell.font = Font(color= colors.BLACK)
        cell.alignment = chosen_alignment
        cell.border = chosen_border

first_row = worksheet1[1]
for cell in first_row:
    cell.font = Font(color= colors.BLUE)
    cell.alignment = Alignment(wrap_text=True,vertical='center',horizontal='center')

worksheet1.column_dimensions['A'].width = 35
worksheet1.column_dimensions['B'].width = 70
worksheet1.row_dimensions[1].height = 40

# Save the xlsx file
workbook.save(xlsx_file_name)

print('saving excel file')

# refresh 
CapellaPlatform.refresh(folder)
