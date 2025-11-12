#
# This script loads the Capella model passed as first argument and imports breakdown of NodePCs from an xlsx file 
# Breakdown of three columns of node PCs with description of the Node PCs in the fourth column, with indentation. 
# Line 1 is the title of column (Level 1, Level 2, Level 3, description). 
# To run it:
#  - enable Developer capabilities if not already done (see documentation in the help menu)
#  - launch the contextual menu "Run As / Run configurations..." on this script
#    - create a new "EASE Script" configuration
#    - define the name of the configuration: "Import_breakdown_Node_PCs_from_xlsx.py" (for instance)
#    - define the Script Source path: "workspace://Python4Capella/sample_scripts/Import_breakdown_Node_PCs_from_xlsx.py"
#    - define the path to the aird file as first argument and the xlsx file as second argument in "Script arguments" area: "/In-Flight Entertainment System/In-Flight Entertainment System.aird" "/Python4Capella/resources/breakdown_Node_PCs.xlsx" (for instance)


# include needed for the Capella modeller API
include('workspace://Python4Capella/simplified_api/capella.py')
if False:
    from simplified_api.capella import *

# include needed for utilities
include('workspace://Python4Capella/utilities/CapellaPlatform.py')
if False:
    from utilities.CapellaPlatform import *

# include needed to read/write xlsx files
from openpyxl import *

#check parameter numbers
if len(argv) != 2:
    # use IFE default values
    aird_path = "/In-Flight Entertainment System/In-Flight Entertainment System.aird"
    xlsx_path = "/Python4Capella/resources/breakdown_Node_PCs.xlsx"
else:
    # Load the Capella model from the first argument of the script
    aird_path = argv[0]
    xlsx_path = argv[1]

model = CapellaModel()
model.open(aird_path)

# gets the SystemEngineering
se = model.get_system_engineering()

# create  a folder in the project
xlsx_file = CapellaPlatform.getWorkspaceFile(xlsx_path)
xlsx_file_name = CapellaPlatform.getAbsolutePath(xlsx_file)

print("Read excel file")

# load the workbook
wb = load_workbook(xlsx_file_name)

# grab the active worksheet
ws = wb.active

# get the PhysicalComponentPkg
pc_pkg = se.get_physical_architecture().get_physical_component_pkg()

# start a transaction to modify the Capella model
model.start_transaction()

i = 1
#indicators about levels of the breakdown
Indicator_new_NPC_of_level_1 = True
Indicator_new_NPC_of_level_2 = True
Indicator_new_NPC_of_level_3 = True

try:
    # create node pc with the list of names in the xlsx file
    for i in range(2,100):
        cell_colum_1 = ws.cell(row = i, column = 1).value
        if cell_colum_1 :
            structure = pc_pkg.get_owned_physical_components()
            # check if name of the NodePC is not already in the structure
            for elem in structure:
                structure_name = elem.get_name()
            if cell_colum_1 not in structure_name:
                Indicator_new_NPC_of_level_1 = True
                # create a NodePC
                npc = NodePC()
                #set its name
                print("* Level 1 - Create a new NodePC with name " + cell_colum_1)
                npc.set_name(cell_colum_1)
                npc_description = ws.cell(row = i, column = 4).value
                if npc_description:
                    npc.set_description(npc_description)
                #add the new NodePC
                pc_pkg.get_owned_physical_components().add(npc)
                line_new_npc = i
                line_npc2 = line_new_npc + 1
                line_npc3 = line_npc2
                cell_colum_2 = ws.cell(row = line_npc2, column = 2).value
                while cell_colum_2 :
                    if cell_colum_2 and (Indicator_new_NPC_of_level_1 is True):
                        structure = pc_pkg.get_owned_physical_components()
                        # check if name of the NodePC is not already in the structure
                        for elem in structure:
                            structure_name = elem.get_name()
                        if cell_colum_2 not in structure_name:
                            Indicator_new_NPC_of_level_2 = True
                            # create a NodePC
                            npc2 = NodePC()
                            #set its name
                            print("* Level 2 - Create a new NodePC with name " + cell_colum_2)
                            npc2.set_name(cell_colum_2)
                            npc2_description = ws.cell(row = line_npc2, column = 4).value
                            if npc2_description:
                                npc2.set_description(npc2_description)
                            #add the new NodePC
                            npc.get_owned_physical_components().add(npc2)
                            line_npc3 = line_npc2 + 1
                            cell_colum_3 = ws.cell(row = line_npc3, column = 3).value
                            while cell_colum_3 :
                                if cell_colum_3 and (Indicator_new_NPC_of_level_1 is True) and (Indicator_new_NPC_of_level_2 is True):
                                    structure = pc_pkg.get_owned_physical_components()
                                    # check if name of the NodePC is not already in the structure
                                    for elem in structure:
                                        structure_name = elem.get_name()
                                    if cell_colum_3 not in structure_name:
                                        Indicator_new_NPC_of_level_3 = True
                                        # create a NodePC
                                        npc3 = NodePC()
                                        #set its name
                                        print("* Level 3 - Create a new NodePC with name " + cell_colum_3)
                                        npc3.set_name(cell_colum_3)
                                        pc3_description = ws.cell(row = line_npc3, column = 4).value
                                        if pc3_description:
                                            npc3.set_description(pc3_description)
                                    #add the new NodePC
                                        npc2.get_owned_physical_components().add(npc3)
                                    
                                    else:
                                        Indicator_new_NPC_of_level_3 = False
                                    
                                    line_npc2 = line_npc3    
                                    line_npc3 = line_npc3 + 1
                                    cell_colum_3 = ws.cell(row = line_npc3, column = 3).value
                                    cell_colum_2 = ws.cell(row = line_npc3+1, column = 2).value
                    else:
                        Indicator_new_NPC_of_level_2 = False
                                            
                    line_npc2 = line_npc2 + 1                        
                    cell_colum_2 = ws.cell(row = line_npc2, column = 2).value                                                 
        else:
            Indicator_new_NPC_of_level_1 = False    
except:
    # if something went wrong we rollback the transaction
    model.rollback_transaction()
    raise
else:
    # if everything is ok we commit the transaction
    model.commit_transaction()

# save the Capella model
model.save()
