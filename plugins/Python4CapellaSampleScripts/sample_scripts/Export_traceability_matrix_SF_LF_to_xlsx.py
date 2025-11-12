'''
This script allows to extract the traceability matrix between System Functions and Logical Functions to xlsx. 
SF without realizing links with LF and LF without realized links with SF are written in red

It will create a folder result in the selected Capella project with the resulting xlsx file.
'''
# To run it:
#  - enable Developer capabilities if not already done (see documentation in the help menu)
#  - you can run this script by launching the contextual menu "Run As / EASE Script..." 
#    on this script. 
#    - By default, the model selected is IFE sample (aird path of the model written below)

#  - you can also run this script according to a configuration (script selected, arguments) 
#    and modify the configuration by launching the contextual menu "Run As / Run configurations..." 
#    on this script. 
#    - create a new "EASE Script" configuration
#    - define the name of the configuration: "Your_Script.py" (for instance)
#    - define the Script Source path: "workspace://Python4Capella/sample_scripts/Your_Script.py"
#

# include needed for the Capella modeller API
include('workspace://Python4Capella/simplified_api/capella.py')
if False:
    from simplified_api.capella import *
    
# include needed for utilities
include('workspace://Python4Capella/utilities/CapellaPlatform.py')
if False:
    from utilities.CapellaPlatform import *
    
# include needed to read/write xlsx files
from openpyxl import *
from openpyxl.styles import Alignment, NamedStyle, Font, colors, Color, PatternFill
from openpyxl.styles.borders import Border, Side


# change this path to execute the script on your model (here is the IFE sample). 
# comment it if you want to use the "Run configuration" instead
aird_path = '/In-Flight Entertainment System/In-Flight Entertainment System.aird'

'''
#Here is the "Run Configuration" part to uncomment if you want to use this functionality :

#check parameter numbers
if len(argv) != 1:
    # use IFE default values
    aird_path = "/In-Flight Entertainment System/In-Flight Entertainment System.aird"
else:
    # Load the Capella model from the first argument of the script
    aird_path = argv[0]
'''


model = CapellaModel()
model.open(aird_path)

# gets the SystemEngineering and print its name
se = model.get_system_engineering()
print('starting export of model ' + se.get_name())

# preparing excel file export
project_name = aird_path[0:(aird_path.index("/", 1) + 1)]
project = CapellaPlatform.getProject(project_name)
folder = CapellaPlatform.getFolder(project, 'results')
xlsx_file_name = CapellaPlatform.getAbsolutePath(folder) + '/' + 'Export_traceability_matrix_SF_LF_to_xlsx.xlsx'
# create a workbook
workbook = Workbook()

# writing excel file header
worksheet = workbook.active
worksheet.title = 'traceability_matrix_SF_LF'

worksheet["A1"] = 'LF name / SF name'

LF_realizing_functions_name = []
all_LF_realizing_functions_name = []

i=2
j=2

Color_font_red = Font(color= "FF0000")

# retrieving elements from the model
all_SF = se.get_all_contents_by_type(SystemFunction)
all_LF = se.get_all_contents_by_type(LogicalFunction)


    
for SF in all_SF:
    i=2
    worksheet.cell(row = 1, column = j).value = SF.get_name()
    LF_realizing_functions = SF.get_realizing_logical_functions()
    LF_realizing_functions_name = []
    for LF in all_LF:
        if LF_realizing_functions:
            for LF_realizing_functions_elem in LF_realizing_functions:
                LF_realizing_functions_name.append(LF_realizing_functions_elem.get_name())
                all_LF_realizing_functions_name.append(LF_realizing_functions_elem.get_name())
            if LF.get_name() in LF_realizing_functions_name:
                worksheet.cell(row = i, column = j).value = 'X'

        else:
            worksheet.cell(row = 1, column = j).font = Color_font_red
        i = i+1
        
    j = j +1

i =2   
for LF in all_LF:
    worksheet["A" + str(i)] = LF.get_name()
    if LF.get_name() not in all_LF_realizing_functions_name:
        worksheet.cell(row = i, column = 1).font = Color_font_red
    i = i+1
    
column_widths = [] 
bd = Side(border_style='medium')
chosen_border = Border(left=bd, top=bd,right=bd, bottom=bd)
chosen_alignment = Alignment(wrap_text=True,vertical='top')


for row in worksheet.iter_rows():

    for cell in row:
        cell.alignment = chosen_alignment
        cell.border = chosen_border
        
# Save the xlsx file
workbook.save(xlsx_file_name)

print('saving excel file')

# refresh 
CapellaPlatform.refresh(folder)
