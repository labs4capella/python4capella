# name                 : Extract a pseudo hierarchy of PA elements to xlsx
# script-type          : Python
# description          : Extract a pseudo hierarchy of PA elements to xlsx
# popup                : enableFor(org.polarsys.capella.core.data.capellacore.CapellaElement)
'''
This script allows to extract a pseudo-hierarchy of elements defined in Physical Architecture starting from the Node PC,
getting the sub-NodePC and deployed BehaviorPV, from the BehaviorPV getting the sub-Behavior PV and allocated functions

It will create a folder result in the selected Capella project with the resulting xlsx file.
'''
# To run it:
#  - enable Developer capabilities if not already done (see documentation in the help menu)
#  - you can run this script by launching the contextual menu "Run As / EASE Script..." 
#    on this script. 
#    - By default, the model selected is IFE sample (aird path of the model written below)

#  - you can also run this script according to a configuration (script selected, arguments) 
#    and modify the configuration by launching the contextual menu "Run As / Run configurations..." 
#    on this script. 
#    - create a new "EASE Script" configuration
#    - define the name of the configuration: "Export_a_pseudo_hierarchy_of_PA_elements_to_xlsx.py" (for instance)
#    - define the Script Source path: "workspace://Python4Capella/sample_scripts/Export_a_pseudo_hierarchy_of_PA_elements_to_xlsx.py"
#

# include needed for the Capella modeller API
include('workspace://Python4Capella/simplified_api/capella.py')
if False:
    from simplified_api.capella import *
    
# include needed for utilities
include('workspace://Python4Capella/utilities/CapellaPlatform.py')
if False:
    from utilities.CapellaPlatform import *
    
# include needed to read/write xlsx files
from openpyxl import *
from openpyxl.formatting.rule import ColorScaleRule,CellIsRule,FormulaRule
from openpyxl.styles import Alignment, NamedStyle, Font, colors, Color, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles.borders import Border, Side

# we define a generic method to retrieve the sub-elements of any kind of element
def getSubElements(ws, i, j, elem):
    ws.cell(row = i, column = j).value = elem.get_name()
    i = i+1
    if (isinstance(elem, NodePC)):
        npc = elem
        #: :type npc: NodePC
        # if we have a NodePC we want the sub-NodePC and deployed BehaviorPC
        for subPC in npc.get_owned_physical_components():
            i = getSubElements(ws, i, j+1, subPC)
        for bpc in npc.get_deployed_behavior_p_cs():
            i = getSubElements(ws, i, j+1, bpc)
    if isinstance(elem, BehaviorPC):
        bpc = elem
        #: :type bpc: BehaviorPC
        # if we have a BehaviorPC we want the sub-BehaviorPC and allocated Functions
        for subPC in bpc.get_owned_physical_components():
            i = getSubElements(ws, i, j+1, subPC)
        for func in bpc.get_allocated_functions():
            i = getSubElements(ws, i, j+1, func)
    # we have nothing more to do if we have a function
    return i

# Retrieve the Element from the current selection and its aird model path
selected_elem = CapellaElement(CapellaPlatform.getFirstSelectedElement())
aird_path = '/'+ CapellaPlatform.getModelPath(selected_elem)

'''
# change this path to execute the script on your model (here is the IFE sample). 
# Uncomment it if you want to use the "Run configuration" instead
aird_path = '/In-Flight Entertainment System/In-Flight Entertainment System.aird'
'''
'''
#Here is the "Run Configuration" part to uncomment if you want to use this functionality :

#check parameter numbers
if len(argv) != 1:
    # use IFE default values
    aird_path = "/In-Flight Entertainment System/In-Flight Entertainment System.aird"
else:
    # Load the Capella model from the first argument of the script
    aird_path = argv[0]
'''

model = CapellaModel()
model.open(aird_path)

# gets the SystemEngineering and print its name
se = model.get_system_engineering()
print('starting export of model ' + se.get_name())

# preparing excel file export
project_name = aird_path[0:(aird_path.index("/", 1) + 1)]
project = CapellaPlatform.getProject(project_name)
folder = CapellaPlatform.getFolder(project, 'results')
xlsx_file_name = CapellaPlatform.getAbsolutePath(folder) + '/' + 'Export_a_pseudo_hierarchy_of_PA_elements_to_xlsx.xlsx'
# create a workbook
workbook = Workbook()

# writing excel file header
worksheet = workbook.active
worksheet.title = 'PA export'
worksheet["A1"] = 'Elem 1'
worksheet["B1"] = 'Elem 2'
worksheet["C1"] = 'Elem 3'
worksheet["D1"] = 'Elem 4'
worksheet["E1"] = 'Elem 5'
worksheet["F1"] = 'Elem 6'
worksheet["G1"] = 'Elem 7'
worksheet["H1"] = 'Elem 8'
worksheet["I1"] = 'Elem 9'

i=2

# retrieving elements from the model
for npc in se.get_physical_architecture().get_physical_system().get_owned_physical_components():
    if (isinstance(npc, NodePC)):
        i = getSubElements(worksheet, i, 1, npc)
        
bd = Side(border_style='medium')
chosen_border = Border(left=bd, top=bd,right=bd, bottom=bd)
chosen_alignment = Alignment(wrap_text=True,vertical='top')


first_row = worksheet[1]
Color_font = Font(color= colors.BLUE)
for cell in first_row:
    cell.font = Color_font
          
for row in worksheet.iter_rows():

    for cell in row:
        cell.alignment = chosen_alignment
        cell.border = chosen_border
        
column_width = 17
worksheet.column_dimensions['A'].width = column_width
worksheet.column_dimensions['B'].width = column_width
worksheet.column_dimensions['C'].width = column_width
worksheet.column_dimensions['D'].width = column_width
worksheet.column_dimensions['E'].width = column_width  
worksheet.column_dimensions['F'].width = column_width  
worksheet.column_dimensions['G'].width = column_width  
worksheet.column_dimensions['H'].width = column_width
worksheet.column_dimensions['I'].width = column_width  
        
# Save the xlsx file
workbook.save(xlsx_file_name)

print('saving excel file')

# refresh 
CapellaPlatform.refresh(folder)
