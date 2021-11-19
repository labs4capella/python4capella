'''
This script allows to export Capabilities and owned Scenarios to Excel
It will create a folder result in the selected Capella project with the resulting xlsx file.
'''
# To run it:
#  - enable Developer capabilities if not already done (see documentation in the help menu)
#  - you can run this script by launching the contextual menu "Run As / EASE Script..." 
#    on this script. 
#    - By default, the model selected is IFE sample (aird path of the model written below)

#  - you can also run this script according to a configuration (script selected, arguments) 
#    and modify the configuration by launching the contextual menu "Run As / Run configurations..." 
#    on this script. 
#    - create a new "EASE Script" configuration
#    - define the name of the configuration: "Export_capabilities_and_owned_scenarios_to_xlsx.py" (for instance)
#    - define the Script Source path: "workspace://Python4Capella/sample_scripts/Export_capabilities_and_owned_scenarios_to_xlsx.py"
#


# include needed for the Capella modeller API
include('workspace://Python4Capella/simplified_api/capella.py')
if False:
    from simplified_api.capella import *
    
 # include needed for utilities
include('workspace://Python4Capella/utilities/CapellaPlatform.py')
if False:
    from utilities.CapellaPlatform import *
    
# include needed to read/write xlsx files
from openpyxl import *

# change this path to execute the script on your model (here is the IFE sample). 
# comment it if you want to use the "Run configuration" instead
aird_path = '/In-Flight Entertainment System/In-Flight Entertainment System.aird'

'''
#Here is the "Run Configuration" part to uncomment if you want to use this functionality :

#check parameter numbers
if len(argv) != 1:
    # use IFE default values
    aird_path = "/In-Flight Entertainment System/In-Flight Entertainment System.aird"
else:
    # Load the Capella model from the first argument of the script
    aird_path = argv[0]
'''

model = CapellaModel()
model.open(aird_path)

# gets the SystemEngineering and print its name
se = model.get_system_engineering()
print('starting export of model ' + se.get_name())

# preparing excel file export
project_name = aird_path[0:(aird_path.index("/", 1) + 1)]
project = CapellaPlatform.getProject(project_name)
folder = CapellaPlatform.getFolder(project, 'results')
xlsx_file_name = CapellaPlatform.getAbsolutePath(folder) + '/' + 'Capability_owned_scenario.xlsx'
# create a workbook
workbook = Workbook()

# writing excel file header
worksheet = workbook.active
worksheet.title = 'Capability and owned scenario'
worksheet["A1"] = 'Capability name'
worksheet["B1"] = 'Owned Scenario name'

# retrieve the list of Capabilities
all_Capabilities = []
all_Capabilities = se.get_all_contents_by_type(Capability)
# now retrieving cpabilities and owned scenarios from model
i = 1

for elem_capability in all_Capabilities:
    i = i + 1
    worksheet.cell(row = i, column = 1).value = elem_capability.get_name()
    involved_scenario = elem_capability.get_owned_scenarios()
    
    for elem_scenario in involved_scenario:
        i = i + 1    
        worksheet.cell(row = i, column = 2).value = elem_scenario.get_name()
        
# Save the xlsx file
workbook.save(xlsx_file_name)

print('saving excel file')

# refresh 
CapellaPlatform.refresh(folder)
