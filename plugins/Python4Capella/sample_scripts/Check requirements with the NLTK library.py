# name                 : Check requirements with the NLTK library
# script-type          : Python
# description          : Check whether the requirements have been defined following good practices inspired by INCOSE's Guide for writing requirements
# popup                : enableFor(org.polarsys.capella.core.data.capellacore.CapellaElement)

'''
The objective of this script is to check if the requirements of a Capella model follow a number of good practices stated in INCOSE's Guide for Writing Requirements.
For that, the script uses the Natural Language ToolKit (NLTK) library for python.

For more information check the library official web page: https://www.nltk.org/

'''
# include needed for the requirement API
include('workspace://Python4Capella/simplified_api/requirement.py')
if False:
    from simplified_api.requirement import *

class NLTK_Processing():
    
    import nltk
    
    requirement = Requirement()
    
    def set_Requirement(self, requirement):
        self.requirement = requirement
    
    def process_requirement(self):
        tokenized_words = self.nltk.word_tokenize(str(self.requirement.get_text()))
        tagged_words = self.nltk.pos_tag(tokenized_words)
        
        stopwords = self.nltk.corpus.stopwords.words("english")
        
        return tokenized_words, tagged_words, stopwords
        
        
class IChecker(): 
    
    def check(self):
        """Check if the requirements comply with the good practice"""
        pass

    def set(self):
        """Set the information in the GoodRequirementChecker"""
        pass

class GoodPracticeChecker_Check_for_keyword_shall(IChecker):
    
    requirement = Requirement()
    keyword = ""
    tokenized_words = []
    
    def set(self, req, keyword, tokenized_words):
        self.requirement = req
        self.keyword = keyword
        self.tokenized_words = tokenized_words
    
    def check(self):
        req = str(self.requirement.get_text())
        try:
            if self.keyword.lower() not in [word.lower() for word in self.tokenized_words]:
                return("Keyword " + "'" + self.keyword + "'" + "not found " + "in req " + str(self.requirement.get_all_attributes()[1].get_value()) + ". It is good practice to use 'shall' when defining requirements.")
        except Exception as e:
            print("Error: " + str(e))
class GoodPracticeChecker_Check_for_repeated_words(IChecker):
    
    import string
    
    requirement = Requirement()
    tokenized_words = []
    english_stopwords = []
    
    def set(self, req, tokenized_words, stopwords):
        self.requirement = req
        self.tokenized_words = tokenized_words
        self.english_stopwords = stopwords
    
    def check(self):
        req = str(self.requirement.get_text())
        
        #Delete punctuation marks
        processed_req = req.translate (str.maketrans ("", "", self.string.punctuation))
        
        #Delete the stopwords
        filtered_words = [word.lower() for word in self.tokenized_words if word.lower() not in self.english_stopwords]
        
        #Check for repeated words in the sentence
        duplicates = [word for word in self.tokenized_words if filtered_words.count(word) > 1]
        unique_duplicates = list(set(duplicates))
        
        if len(duplicates)>0:
            return(f"Be careful, repeated words found: {str(unique_duplicates)}. It is good practice not to repeat words." )

class GoodPracticeChecker_Check_use_of_indefinite_article_the_instead_of_a(IChecker):
    
    requirement = Requirement()
    tagged_words = []
    
    def set(self, req, tagged_words):
        self.requirement = req
        self.tagged_words = tagged_words
    
    def check(self):
        for tagged_word in self.tagged_words:
        #If there is a DT different than "the" that is not followed by a noun (or adjective + noun) show error message
            if tagged_word[1] == "DT" and tagged_word[0].lower() != "the":     
                return (f"The use of the indefinite article '{tagged_word[0]}' was detected in requirement {str(self.requirement.get_all_attributes()[1].get_value())}. Good practice recommends using 'the' instead of 'a/an' when defining requirements.")


class GoodPracticeChecker_Check_for_use_of_vague_terms(IChecker):
    
    requirement = Requirement()
    tokenized_words = []
    vague_qualifications = []
    vague_adjectives = []
    vague_adverbs = []
    
    def set(self, req, tokenized_words, vague_qualifications, vague_adjectives, vague_adverbs):
        self.requirement = req
        self.tokenized_words = tokenized_words
        self.vague_qualifications = vague_qualifications
        self.vague_adjectives = vague_adjectives
        self.vague_adverbs = vague_adverbs

    def check(self):

        vague_words = []
        
        for word in self.tokenized_words:
            if word.lower() in (self.vague_qualifications or self.vague_adjectives or self.vague_adverbs):
                vague_words.append(word)
        if len(vague_words)>0:
            return(f"Vague words ({str(vague_words)}) identified in requirement {str(self.requirement.get_all_attributes()[1].get_value())}. Good practice recommends its use when developing requirements.") 

class GoodPracticeChecker_Check_for_use_of_not(IChecker):
    
    requirement = Requirement()
    tagged_words = []
    
    def set(self, req, tagged_words):
        self.requirement = req
        self.requirement = tagged_words
    
    def check(self):

        for tagged_word in self.tagged_words:
            if tagged_word[1] == "RB" and tagged_word[0].lower() in ["not", "n't"]:
                return(f"Use of 'Not' words identified in requirement {str(self.requirement.get_all_attributes()[1].get_value())}. Good practice recommends not recommend the use of negative terms when developing requirements.")

class GoodPracticeChecker_Check_for_use_of_slash(IChecker):
    
    requirement = Requirement()
    tagged_words = []
    
    def set(self, req, tagged_words):
        self.requirement = req
        self.tagged_words = tagged_words
    
    def check(self):
        contains_slash = False
    
        for tagged_word in self.tagged_words:
            if tagged_word[0].__contains__(str('/')) and tagged_word[1] != "CD": 
                    contains_slash = True
    
        if contains_slash:
            return(f"The use of '/' was detected in requirement {str(self.requirement.get_all_attributes()[1].get_value())}. Good practice does not recommend it's use when developing requirements.")

class GoodPracticeChecker_Check_for_use_of_combinator_words(IChecker):
    
    requirement = Requirement()
    tokenized_words = []
    combinator_words = []
    
    def set(self, req, tokenized_words, combinators):
        self.requirement = req
        self.tokenized_words = tokenized_words
        self.combinator_words = combinators

    def check(self):
        combinators_found = []
        for i in range(len(self.tokenized_words)):
            if (self.tokenized_words[i].lower()) in self.combinator_words:
                combinators_found.append(self.tokenized_words[i])
                
                #check if 2-word combinator
            if i+1 < len(self.tokenized_words):
                if (self.tokenized_words[i].lower() + " " + self.tokenized_words[i+1].lower()) in self.combinator_words:
                    combinators_found.append(self.tokenized_words[i] + " " + self.tokenized_words[i+1])
                    i = i + 1
                            
            # check if 3-word combination
            if i+2 < len(self.tokenized_words):
                if (self.tokenized_words[i].lower() + " " + self.tokenized_words[i+1].lower() + " " + self.tokenized_words[i+2].lower()) in self.combinator_words:
                    combinators_found.append(self.tokenized_words[i] + " " + self.tokenized_words[i+1] + " " + self.tokenized_words[i+2])
                    i = i + 2
    
        if len(combinators_found)>0: 
            return(f"Combinator words ({str(combinators_found)}) detected in requirement {str(self.requirement.get_all_attributes()[1].get_value())}. Good practice does not recommend it's use for developing requirements.") 
    
class GoodPracticeChecker_Check_for_use_of_parentheses(IChecker):
    
    requirement = Requirement()
    tokenized_words = []
    
    def set(self, req, tokenized_words):
        self.requirement = req
        self.tokenized_words = tokenized_words
    
    def check(self):
        if "(" in self.tokenized_words:
            if ")" in self.tokenized_words:
                return(f"Use of parentheses and/or brackets detected in requirement {str(self.requirement.get_all_attributes()[1].get_value())}. Good practice does not recommend the use of parentheses and /or brackets containing subordinate text.")
        
        elif "[" in self.tokenized_words:
            if "]" in self.tokenized_words:
                return(f"The use of parentheses and/or brackets was detected in requirement {str(self.requirement.get_all_attributes()[1].get_value())}. Good practice does not recommend the use of parentheses and /or brackets containing subordinate text.")
     
class GoodPracticeChecker_Check_for_use_of_pronouns_and_indefinite_pronouns(IChecker):
    
    requirement = Requirement()
    tagged_words = []
    indefinitie_pronouns = []
    
    def set(self, req, tagged_words, indefinite_pronouns):
        self.requirement = req
        self.tagged_words = tagged_words
        self.indefinitie_pronouns = indefinite_pronouns
    
    def check(self):
        
        # PRP: Pronouns
        # DT: some indefinite pronouns

        found_pronouns = []
    
        for word in self.tagged_words:
            if word[1] != "PRP" and word[1] != "DT" and word[0].lower() not in self.indefinitie_pronouns:
                pass
            else:
                if word[0].lower() != "the" and word[0] not in found_pronouns:
                    found_pronouns.append(word[0])
        
        if len(found_pronouns)>0:
            return(f"The use of pronouns and or indefinite pronouns was detected: {found_pronouns} in requirement {str(self.requirement.get_all_attributes()[1].get_value())}. Good practices recommend avoiding the use of pronouns and indefinite pronouns.")
        
class GoodPracticeChecker_Check_for_unachievable_absolutes(IChecker):
    
    requirement = Requirement()
    tokenized_words = []
    absolute_words = []
    
    
    def set(self, req, tokenized_words, absolute_words):
        self.requirement = req
        self.tokenized_words = tokenized_words
        self.absolute_words = absolute_words
    
    def check(self):
        found_words = []

        for word in self.tokenized_words:
            if word[0].lower() in self.absolute_words:
                found_words.append(word[0]) 
        
        if found_words:
            return(f"The use of unachievable absolutes has been detected: {found_words} in requirement {str(self.requirement.get_text())}. Good practice recommends avoiding the use of unachievable absolutes.")                 
            
class GoodPracticeRepeatedRequirements(IChecker):
    
    reqs = []
    
    def set(self, reqs):
        self.reqs = reqs
    
    def check(self):
        list_reqs = []
        list_reqs_string = []
        list_ids = []
        
        for req in self.reqs:
            list_reqs.append(req)
            list_reqs_string.append(str(req.get_text()))
        
        for i in range(len(list_reqs_string)):
            for j in range(len(list_reqs_string)):
                if list_reqs_string[i] == list_reqs_string[j] and i!=j:
                    if str(list_reqs[i].get_all_attributes()[1].get_value()) not in list_ids:
                        list_ids.append(str(list_reqs[i].get_all_attributes()[1].get_value()))
                    if str(list_reqs[j].get_all_attributes()[1].get_value()) not in list_ids:
                        list_ids.append(str(list_reqs[j].get_all_attributes()[1].get_value()))
        
        if len(list_ids)>0:
            return(f"The following requirements are repeated: {list_ids}. Good practice recommends expressing each requirement only once.")
        

class GoodPracticeChecker_Check_for_measurable_target_performances(IChecker):
    
    requirement = Requirement()
    tokenized_words = []
    ambiguous_target_performances = []
    ambiguous_expressions = []
    
    def set(self, req, tokenized_words, ambiguous_target_performances, ambiguous_expressions):
        self.requirement = req
        self.tokenized_words = tokenized_words
        self.ambiguous_target_performances = ambiguous_target_performances
        self.ambiguous_expressions = ambiguous_expressions
    
    def check(self):
        list_found = []
        
        for word in self.tokenized_words:
            if word in self.ambiguous_target_performances:
                list_found.append(word)
        
        for word in self.ambiguous_expressions:
            if word in (str(self.requirement.get_text())).lower():
                list_found.append(word)
        
        if len(list_found)>0:
            return(f"The use of the following ambiguous terms have been found: {list_found} in requirement {str(self.requirement.get_all_attributes()[1].get_value())}. Good practice recommends the use of non-ambiguous terms.")
        
class GoodPracticeChecker_Check_for_measurable_target_timing(IChecker):
    
    requirement = Requirement()
    tokenized_words = []
    ambiguous_timing = []

    def set(self, requirement, tokenized_words, ambiguous_timing):
        self.requirement = requirement
        self.tokenized_words = tokenized_words
        self.ambiguous_timing = ambiguous_timing
    
    def check(self):
        
        list_found = []
        
        for word in self.tokenized_words:
            if word in self.ambiguous_timing:
                list_found.append(word)
        
        if len(list_found)>0:
            return(f"The use of the following ambiguous temporal dependencies have been found: {list_found} in requirement {str(self.requirement.get_all_attributes()[1].get_value())}. Good practice recommends the use of non-ambiguous temporal dependencies.")
        
class Exporter():
    
    import openpyxl
    
    feedback_messages = []
    
    def set_messages(self, dictionary):
        self.feedback_messages = dictionary
    
    def print_messages(self):
        print("---------------------------------------------------------------------------")
        
        for i, req in enumerate(self.feedback_messages.keys()):
            print(f"{i+1} - {req}")
            messages = self.feedback_messages.get(req)
            
            for j, message in enumerate(messages):
                if message:
                    print(f"\t - {message}")
            
            print()

    
    def export_to_excel(self, path):
        # preparing excel file export
        project_name = path[0:(path.index("/", 1) + 1)]
        project = CapellaPlatform.getProject(project_name)
        folder = CapellaPlatform.getFolder(project, 'results')
        xlsx_file_name = CapellaPlatform.getAbsolutePath(folder) + '/' + 'Requirements_good_practices.xlsx'
        # create a workbook
        workbook = Workbook()

        # writing excel file header
        worksheet = workbook.active
        worksheet.title = 'REQs Good Practices'
        worksheet["A1"] = 'Requirement'
        worksheet["B1"] = 'Good practices'
        
        i = 2
        for req in self.feedback_messages.keys():
            worksheet.cell(row = i, column = 1).value = req
            
            messages = self.feedback_messages.get(req)
            if messages:
                for message in messages:
                    if message != None:
                        worksheet.cell(row = i, column = 2).value = message
                        i += 1
            else:
                i += 1
    
        # Save the xlsx file
        workbook.save(xlsx_file_name)
        
        # refresh 
        CapellaPlatform.refresh(folder)

class RequirementsProcessing():
    
    path = ""
    
    def set_path(self, path):
        
        self.path = path
    
    def get_requirements(self):        
        res = []
        
        model = CapellaModel()
        model.open(self.path)
        
        se = model.get_system_engineering()
        
        for req in se.get_all_contents_by_type(Requirement):
            if str(req.get_text()) != "None":
                res.append(req)

        return res
                

class Requirements_Checker:
    
    list_of_good_practices = []
    reqs = []
    dict_forbidden_words = []
    path = ""
    dict_messages = {}
    
    nltk_processor = NLTK_Processing()
    exporter = Exporter()
    requirements_processor = RequirementsProcessing()
    
    def __init__(self, path):
        
        # set path
        self.path = path
        
        # get requirements:
        self.requirements_processor.set_path(path)
        self.reqs = self.requirements_processor.get_requirements()
        
        # instantiate Good Practice classes
        self.list_of_good_practices.append(GoodPracticeChecker_Check_for_keyword_shall())
        self.list_of_good_practices.append(GoodPracticeChecker_Check_for_repeated_words())
        self.list_of_good_practices.append(GoodPracticeChecker_Check_use_of_indefinite_article_the_instead_of_a())
        self.list_of_good_practices.append(GoodPracticeChecker_Check_for_use_of_vague_terms())
        self.list_of_good_practices.append(GoodPracticeChecker_Check_for_use_of_not())
        self.list_of_good_practices.append(GoodPracticeChecker_Check_for_use_of_slash())
        self.list_of_good_practices.append(GoodPracticeChecker_Check_for_use_of_combinator_words())
        self.list_of_good_practices.append(GoodPracticeChecker_Check_for_use_of_parentheses())
        self.list_of_good_practices.append(GoodPracticeChecker_Check_for_use_of_pronouns_and_indefinite_pronouns())
        self.list_of_good_practices.append(GoodPracticeChecker_Check_for_unachievable_absolutes())
        self.list_of_good_practices.append(GoodPracticeChecker_Check_for_measurable_target_performances())
        self.list_of_good_practices.append(GoodPracticeChecker_Check_for_measurable_target_timing())
        
        # define list of forbidden words
        
        self.dict_forbidden_words = {
            "Vague qualifications": ["some", "only", "allowable", "several", "many", "lot", "few", "almost", "always", "nearly", "about", "approximate"],
            "Vague adjectives": ["ancillary", "relevant", "routine", "common", "generic", "significant", "flexible", "expandable", "typical", "sufficient", "adequate", "appropriate", "efficient", "effective", "proficient", "reasonable", "customary"],
            "Vague adverbs": ["usually", "approximately", "sufficiently", "typically"], 
            "Combinators": ["then", "unless", "but", "as well as", "but also", "however", "whether", "meanwhile", "whereas", "otherwise", "although", "even though", "even if", "in case", "in spite of", "despite", "so that", "whatever", "whenever"],
            "Absolute words": ["all", "always", "never", "100%", "100 %", ],
            "Indefinite pronouns": ["all", "another", "any", "anybody", "anything", "both", "each", "either", "every", "everybody", "everyone", "everything", "few", "many", "most", "neither", "no one", "nobody", "none", "one", "several", "some", "somebody", "someone", "something", "such"],
            "Measurable target words": ["prompt", "routine", "maximum", "minimum", "optimum", "optimal", "nominal"],
            "Ambiguous expressions": ["easy to use", "close quickly", "medium-sized", "high-speed", "best practices", "user-friendly"],
            "Ambiguous timing": ["eventually", "until", "before", "when", "after", "as", "once", "earliest", "lastest", "instantaneous", "simultaneous", "while"]
            }

    def general_check(self):
        
        for req_index, req in enumerate(self.reqs):
            
            # process requirement using NLTK:
            self.nltk_processor.set_Requirement(req)
            tokenized_words, tagged_words, english_stopwords = self.nltk_processor.process_requirement()
            
            # update the information contained in each good practice checker
            self.list_of_good_practices[0].set(req, "Shall", tokenized_words)
            self.list_of_good_practices[1].set(req, tokenized_words, english_stopwords)
            self.list_of_good_practices[2].set(req, tagged_words)
            self.list_of_good_practices[3].set(req, tokenized_words, self.dict_forbidden_words["Vague qualifications"], self.dict_forbidden_words["Vague adjectives"], self.dict_forbidden_words["Vague adverbs"])
            self.list_of_good_practices[4].set(req, tagged_words)
            self.list_of_good_practices[5].set(req, tagged_words)
            self.list_of_good_practices[6].set(req, tokenized_words, self.dict_forbidden_words["Combinators"])
            self.list_of_good_practices[7].set(req, tokenized_words)
            self.list_of_good_practices[8].set(req, tagged_words, self.dict_forbidden_words["Indefinite pronouns"])
            self.list_of_good_practices[9].set(req, tokenized_words, self.dict_forbidden_words["Absolute words"])
            self.list_of_good_practices[10].set(req, tokenized_words, self.dict_forbidden_words["Measurable target words"], self.dict_forbidden_words["Ambiguous expressions"])
            self.list_of_good_practices[11].set(req, tokenized_words, self.dict_forbidden_words["Ambiguous timing"])
            
            feedback_messages = []
            for good_practice in self.list_of_good_practices:
                feedback = good_practice.check()
                feedback_messages.append(feedback)
            
            self.dict_messages[str(req.get_text())] = feedback_messages
            
        
        goodPracticeRepeatedRequirements = GoodPracticeRepeatedRequirements()
        goodPracticeRepeatedRequirements.set(self.reqs)
        feedback = goodPracticeRepeatedRequirements.check()
        
        self.dict_messages["Other"] = [feedback]
        
    def export_to_console(self):
        self.exporter.set_messages(self.dict_messages)
        self.exporter.print_messages()
    
    def export_to_excel(self):
        self.exporter.set_messages(self.dict_messages)
        self.exporter.export_to_excel(self.path)
    

aird_path =  '/In-Flight Entertainment System/In-Flight Entertainment System.aird'

r = Requirements_Checker(aird_path)
r.general_check()
r.export_to_console()
r.export_to_excel()
