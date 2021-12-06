# name                 : Export missions, capabilities, FC and progress statuses
# script-type          : Python
# description          : Export missions, capabilities, FC and progress statuses
# popup                : enableFor(org.polarsys.capella.core.data.capellacore.CapellaElement)


'''
This script allows to export missions, exploited capabilities and involved Functional Chains to Excel
It will create a folder result in the selected Capella project with the resulting xlsx file.
'''

# To run it:
#  - enable Developer capabilities if not already done (see documentation in the help menu)
#  - you can run this script by launching the contextual menu "Run As / EASE Script..." 
#    on this script. 
#    - By default, the model selected is IFE sample (aird path of the model written below)

#  - you can also run this script according to a configuration (script selected, arguments) 
#    and modify the configuration by launching the contextual menu "Run As / Run configurations..." 
#    on this script. 
#    - create a new "EASE Script" configuration
#    - define the name of the configuration: "Export_Missions_capabilities_FC_status.py" (for instance)
#    - define the Script Source path: "workspace://Python4Capella/sample_scripts/Export_Missions_capabilities_FC_status.py"
#


# include needed for the Capella modeller API
include('workspace://Python4Capella/simplified_api/capella.py')
if False:
    from simplified_api.capella import *
    
# include needed for the PVMT API
include('workspace://Python4Capella/simplified_api/pvmt.py')
if False:
    from simplified_api.pvmt import *
    
 # include needed for utilities
include('workspace://Python4Capella/utilities/CapellaPlatform.py')
if False:
    from utilities.CapellaPlatform import *
    
# include needed to read/write xlsx files
from openpyxl import *
from openpyxl.formatting.rule import ColorScaleRule,CellIsRule,FormulaRule
from openpyxl.styles import Alignment, NamedStyle, Font, colors, Color, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles.borders import Border, Side

# Retrieve the Element from the current selection and its aird model path
selected_elem = CapellaElement(CapellaPlatform.getFirstSelectedElement())
aird_path = '/'+ CapellaPlatform.getModelPath(selected_elem)

'''
# change this path to execute the script on your model (here is the IFE sample). 
# Uncomment it if you want to use the "Run configuration" instead
aird_path = '/In-Flight Entertainment System/In-Flight Entertainment System.aird'
'''
'''
#Here is the "Run Configuration" part to uncomment if you want to use this functionality :
#check parameter numbers
if len(argv) != 1:
    # use IFE default values
    aird_path = "/In-Flight Entertainment System/In-Flight Entertainment System.aird"
else:
    # Load the Capella model from the first argument of the script
    aird_path = argv[0]
'''

model = CapellaModel()
model.open(aird_path)

# gets the SystemEngineering and print its name
se = model.get_system_engineering()
print('starting export of model ' + se.get_name())

# preparing excel file export
project_name = aird_path[0:(aird_path.index("/", 1) + 1)]
project = CapellaPlatform.getProject(project_name)
folder = CapellaPlatform.getFolder(project, 'results')
xlsx_file_name = CapellaPlatform.getAbsolutePath(folder) + '/' + 'Mission_Capability_FC_and_progress_status.xlsx'
# create a workbook
workbook = Workbook()

# writing excel file header
worksheet1 = workbook.active
worksheet1.title = 'Mission, Capability, statuses'
worksheet1["A1"] = 'Mission name'
worksheet1["B1"] = 'Exploited Capability name'
worksheet1["C1"] = 'Percentage Designed'
worksheet1["D1"] = 'Percentage Developed'
worksheet1["E1"] = 'Percentage Validated'

# retrieve the list of Missions and PVs
all_Missions = []
allPVs = []

all_Missions = se.get_all_contents_by_type(Mission)

# now retrieving Missions, exploited Capabilities and progress statuses defined with PVMT add on from model
i = 1
j = 2

for elem_Mission in all_Missions:
    i = i + 1
    worksheet1.cell(row = i, column = 1).value = elem_Mission.get_name()
    exploited_capability = elem_Mission.get_exploited_capabilities()
    
    for elem_capability in exploited_capability: 
        i = i + 1
        j = 2    
        worksheet1.cell(row = i, column = 2).value = elem_capability.get_name()
        for pvName in PVMT.get_p_v_names(elem_capability):
            if pvName not in allPVs:
                allPVs.append(pvName)
        for pvName in allPVs:
            if pvName == 'Percentage_Designed':
                j = j + 1
                worksheet1.cell(row = i, column = j).value = int(PVMT.get_p_v_value(elem_capability, pvName))
            if pvName == 'Percentage_Developed':
                j = j + 1
                worksheet1.cell(row = i, column = j).value = int(PVMT.get_p_v_value(elem_capability, pvName))       
            if pvName == 'Percentage_Validated':
                j = j + 1
                worksheet1.cell(row = i, column = j).value = int(PVMT.get_p_v_value(elem_capability, pvName))        

 
column_widths = [] 
bd = Side(border_style='medium')
chosen_border = Border(left=bd, top=bd,right=bd, bottom=bd)
chosen_alignment = Alignment(wrap_text=True,vertical='top')


first_row = worksheet1[1]
Color_font = Font(color= colors.BLUE)
for cell in first_row:
    cell.font = Color_font
          
for row in worksheet1.iter_rows():

    for cell in row:
        cell.alignment = chosen_alignment
        cell.border = chosen_border

worksheet1.conditional_formatting.add('C1:E100', ColorScaleRule(start_type='min', start_value=0,start_color ='AA0000', end_type ='max', end_value=100, end_color='F7FF3C'))

column_width = 30
worksheet1.column_dimensions['A'].width = column_width
worksheet1.column_dimensions['B'].width = column_width
column_width = 20
worksheet1.column_dimensions['C'].width = column_width
worksheet1.column_dimensions['D'].width = column_width
worksheet1.column_dimensions['E'].width = column_width

# writing excel file header
worksheet2 = workbook.create_sheet('Mission, capabilities, FC', 2)
worksheet2["A1"] = 'Mission name'
worksheet2["B1"] = 'Exploited Capability name'
worksheet2["C1"] = 'Involved Functional Chain name'


# retrieve the list of Missions and PVs
all_Missions = []
all_Missions = se.get_all_contents_by_type(Mission)

# now retrieving Missions, exploited Capabilities and involved Functional Chains from model
i = 1

for elem_Mission in all_Missions:
    i = i + 1
    worksheet2.cell(row = i, column = 1).value = elem_Mission.get_name()
    exploited_capability = elem_Mission.get_exploited_capabilities()
    
    for elem_capability in exploited_capability: 
        i = i + 1    
        worksheet2.cell(row = i, column = 2).value = elem_capability.get_name()
        involved_FC = elem_capability.get_involved_functional_chains()
        
        for elem_involved_FC in involved_FC: 
            i = i + 1    
            worksheet2.cell(row = i, column = 3).value = elem_involved_FC.get_name()
    
column_widths = []            
first_row = worksheet2[1]
Color_font = Font(color= colors.BLUE)
for cell in first_row:
    cell.font = Color_font
          
for row in worksheet2.iter_rows():

    for cell in row:
        cell.alignment = chosen_alignment
        cell.border = chosen_border

column_width = 30
worksheet2.column_dimensions['A'].width = column_width
worksheet2.column_dimensions['B'].width = column_width
worksheet2.column_dimensions['C'].width = column_width
worksheet2.column_dimensions['D'].width = column_width
                   
# Save the xlsx file
workbook.save(xlsx_file_name)

print('saving excel file')

# refresh 
CapellaPlatform.refresh(folder)
