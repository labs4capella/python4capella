'''
This script allow to extract Property Values defined with PVMT associated to elements (in this case SystemFunctions)
'''

# include needed for the Capella modeller API
include('workspace://Python4Capella/simplified_api/capella.py')
if False:
    from simplified_api.capella import *
    
# include needed for the PVMT API
include('workspace://Python4Capella/simplified_api/pvmt.py')
if False:
    from simplified_api.pvmt import *
    
# include needed for utilities
include('workspace://Python4Capella/utilities/CapellaPlatform.py')
if False:
    from utilities.CapellaPlatform import *
    
# include needed to read/write xlsx files
from openpyxl import *

# change this path to execute the script on your model
aird_path = '/In-Flight Entertainment System/In-Flight Entertainment System.aird'
model = CapellaModel()
model.open(aird_path)

# gets the SystemEngineering and print its name
se = model.get_system_engineering()
print('starting export of model ' + se.get_name())

# preparing excel file export
project_name = aird_path[0:(aird_path.index("/", 1) + 1)]
project = CapellaPlatform.getProject(project_name)
folder = CapellaPlatform.getFolder(project, 'results')
CapellaPlatform.getAbsolutePath(folder)
xlsx_file_name = CapellaPlatform.getAbsolutePath(folder) + '/' + 'Need3b.xlsx'
# create a workbook
workbook = Workbook()


# writing excel file header
worksheet = workbook.active
worksheet.title = 'PV export on Functions'
worksheet["A1"] = 'System Function name'

# retrieve the list of PV to write the header
allPVs = []

# change SystemFunction by another type to retrieve PV for other elements
allSF = se.get_all_contents_by_type(SystemFunction)

for sf in allSF:
    for pvName in PVMT.get_p_v_names(sf):
        if pvName not in allPVs:
            allPVs.append(pvName)

# writing header for all PV found
j = 2
for pvName in allPVs:
    worksheet.cell(row = 1, column = j).value = pvName
    j = j + 1

# now retrieving elements from model
i = 2

for sf in allSF:
    worksheet.cell(row = i, column = 1).value = sf.get_name()
    j = 2
    for pvName in allPVs:
        worksheet.cell(row = i, column = j).value = PVMT.get_p_v_value(sf, pvName)
        j = j + 1
    i = i + 1

# Save the xlsx file
workbook.save(xlsx_file_name)

print('saving excel file')

# refresh 
CapellaPlatform.refresh(folder)